import glob
import json
import os
from collections import defaultdict

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, PatternFill, Side
from PIL import Image as Image2


def fetch_nfts_by_creator(creator_id):
    single_creator_created_nfts = []
    current_page = 1
    headers = {
        'Content-Type': 'application/json'
    }

    while 1:
        response = requests.get(
            f'https://www.oursong.com/api/open-api/user/{creator_id}/created-vibe-list?api_key={API_KEY}&page={current_page}&per_page=100',
            headers=headers
        )
        
        if response.status_code != 200:
            print(f"Error fetching data for creator {creator_id}: {response.status_code}")
            break

        result = response.json()
        single_creator_created_nfts.extend([nft['id'] for nft in result.get('list', [])])

        if not result.get('has_more_page', False):
            break

        current_page = result.get('current_page', current_page) + 1

    print(len(single_creator_created_nfts))
    return single_creator_created_nfts


def fetch_single_nft_info(vibe_id):
    headers = {
        'Content-Type': 'application/json'
    }
    
    response = requests.get(
        f'https://www.oursong.com/api/open-api/vibe/{vibe_id}/profile?api_key={API_KEY}',
        headers=headers
    )
    
    if response.status_code != 200:
        print(f"Error fetching data for creator {vibe_id}: {response.status_code}")
        
    result = response.json()
    return result


def fetch_single_nft_holders(vibe_id):
    single_nft_holders = []
    additional_nft_data = {}
    current_page = 1
    
    while 1:
        headers = {
            'Content-Type': 'application/json'
        }
        response = requests.get(
            f'https://www.oursong.com/api/open-api/vibe/{vibe_id}/holder-list?api_key={API_KEY}&page={current_page}&per_page=100',
            headers=headers
        )
        
        if response.status_code != 200:
            print(f"Error fetching data for NFT holders {vibe_id}: {response.status_code}")
            break
        
        result = response.json()

        if current_page == 1: additional_nft_data = result.get('song_project', {})

        single_nft_holders.extend(result.get('holder_list', []))
        
        if not result.get('has_more_page', False):
            break

        current_page = result.get('current_page', current_page) + 1

    return (additional_nft_data, single_nft_holders)



# ----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

API_KEY = os.getenv('API_KEY')

def download_creator_data_as_json(creators):
    data = defaultdict(dict)


    for creator_index, creator_id in enumerate(creators):
        creators_created_nft_ids = fetch_nfts_by_creator(creator_id)
        for nft_index, nft_id in enumerate(creators_created_nft_ids):
            single_nft_info = fetch_single_nft_info(nft_id)
            del single_nft_info['issuer']
            (additional_nft_data, single_nft_holders) = fetch_single_nft_holders(nft_id)
            single_nft_info['token_spec'] = additional_nft_data['token_spec']
            single_nft_info['contract_address'] = additional_nft_data['contract_address']
            data[creator_id][nft_id] = { 'info': single_nft_info, 'holders': single_nft_holders }

            print(f"Processing creator {creator_index + 1}/{len(creators)}: {creator_id}\nProcessing NFT {nft_index + 1}/{len(creators_created_nft_ids)}: {nft_id}")


    with open('./data.json', 'w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)
    print('done')


def xlsx_converter():
    with open('./data.json', 'r', encoding='utf-8') as file:
        data = json.load(file)

    wb = Workbook()
    medium = Side(style='medium')
    border = Border(left=medium, right=medium, top=medium, bottom=medium)
    cursor_position = 1
    c_blue = PatternFill(
        start_color='9dc2fa',
        end_color='9dc2fa',
        fill_type='solid'
    )
    c_orange = PatternFill(
        start_color='ff9369',
        end_color='ff9369',
        fill_type='solid'
    )
    alignment = Alignment(horizontal='center', vertical='center')
    ws = wb.active
    wb.remove(ws)

    for creator_index, creator_id in enumerate(data):
        print(creator_index, creator_id)
        ws = wb.create_sheet(title=creator_id)

        for nft_index, nft in enumerate(data[creator_id]):
            info = data[creator_id][nft]['info']
            holders = data[creator_id][nft]['holders']

            # ---------- basic info, first 2 blue rows ----------
            ws[f'A{cursor_position}'] = '編號 ID'
            ws[f'A{cursor_position + 1}'] = info['id']
            ws.column_dimensions['A'].width = 14.5

            ws[f'B{cursor_position}'] = '標題 Titile'
            ws[f'B{cursor_position + 1}'] = info['title']
            ws.column_dimensions['B'].width = 14.5

            ws[f'C{cursor_position}'] = '名稱 Name'
            ws[f'C{cursor_position + 1}'] = info['name']
            ws.column_dimensions['C'].width = 14.5

            ws[f'D{cursor_position}'] = '描述 Description'
            ws[f'D{cursor_position + 1}'] = info['description']
            ws.column_dimensions['D'].width = 30

            ws[f'E{cursor_position}'] = '圖片 Image'
            response = requests.get(info['cover_image'])
            fn = f'tmp_image/{creator_id}_{nft_index}.jpg'
            with open(fn, 'wb') as file:
                file.write(response.content)
            image = Image2.open(fn)
            new_width = 100
            width_percent = (new_width / float(image.size[0]))
            new_height = int((float(image.size[1]) * float(width_percent)))
            resized_image = image.resize((new_width, new_height), Image2.Resampling.LANCZOS)
            resized_image = resized_image.convert('RGB')
            resized_image.save(fn)
            ws.add_image(Image(fn), f'E{cursor_position + 1}')
            ws.column_dimensions['E'].width = 14.5

            ws[f'F{cursor_position}'] = '媒體類型 Content Type'
            ws[f'F{cursor_position + 1}'] = info['content_type']
            ws.column_dimensions['F'].width = 20.5

            ws[f'G{cursor_position}'] = '建立時間 Created At'
            ws[f'G{cursor_position + 1}'] = info['created_at']
            ws.column_dimensions['G'].width = 20

            ws[f'H{cursor_position}'] = '合約類型 Token Spec'
            ws[f'H{cursor_position + 1}'] = info['token_spec']
            ws.column_dimensions['H'].width = 20

            ws[f'I{cursor_position}'] = '合約地址 Contract Address'
            ws[f'I{cursor_position + 1}'] = info['contract_address']
            ws.column_dimensions['I'].width = 44

            ws.row_dimensions[cursor_position + 1].height = 130

            for row in ws.iter_rows(
                min_row=cursor_position,
                max_row=cursor_position + 1,
                min_col=1,
                max_col=9
            ):
                for cell in row:
                    cell.fill = c_blue
                    cell.border = border
                    cell.alignment = alignment
            cursor_position += 2
            # ---------- basic info, first 2 blue rows ----------



            # ---------- holders, next n white row ----------
            ws[f'A{cursor_position}'] = '使用者內部編號 UUID'
            for row in ws.iter_rows(
                min_row=cursor_position,
                min_col=1,
                max_col=9
            ):
                for cell in row:
                    cell.fill = c_orange
                    cell.border = border
                    cell.alignment = alignment
            ws.merge_cells(f'A{cursor_position}:C{cursor_position}')

            ws[f'D{cursor_position}'] = '使用者編號 ID'
            ws[f'E{cursor_position}'] = '使用者顯示名稱 Name'
            ws[f'G{cursor_position}'] = '使用者名稱 Username'
            ws[f'H{cursor_position}'] = '持有數量 Owned Amount'
            ws.merge_cells(f'E{cursor_position}:F{cursor_position}')
            ws.merge_cells(f'H{cursor_position}:I{cursor_position}')


            cursor_position += 1

            for holder in holders:
                print(f"Processing creator {creator_index + 1}/{len(data)}: {creator_id}\nProcessing NFT {nft_index + 1}/{len(data[creator_id])}: {nft}")
                ws[f'A{cursor_position}'] = holder['uuid']
                ws[f'D{cursor_position}'] = holder['id']
                ws[f'E{cursor_position}'] = holder['name']
                ws[f'G{cursor_position}'] = holder['username']
                ws[f'H{cursor_position}'] = holder['owned_amount']

                for row in ws.iter_rows(
                    min_row=cursor_position,
                    min_col=1,
                    max_col=9
                ):
                    ws.merge_cells(f'A{cursor_position}:C{cursor_position}')
                    ws.merge_cells(f'E{cursor_position}:F{cursor_position}')
                    ws.merge_cells(f'H{cursor_position}:I{cursor_position}')
                    for cell in row:
                        cell.fill = c_orange
                        cell.border = border
                        cell.alignment = alignment

                cursor_position += 1
            
            cursor_position += 2
            # ---------- holders, next n white row ----------
            
    
    wb.save('data.xlsx')

    image_files = glob.glob(os.path.join('tmp_image', '*.jpg')) + glob.glob(os.path.join('tmp_image', '*.png'))
    for image_file in image_files:
        try:
            os.remove(image_file)
            print(f"已刪除檔案: {image_file}")
        except Exception as e:
            print(f"無法刪除檔案 {image_file}: {e}")


